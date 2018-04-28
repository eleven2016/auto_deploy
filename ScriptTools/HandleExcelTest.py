import xlrd
import openpyxl


if __name__ == '__main__':
    source_file_path = "D:\\workspace_test\\AutoDeploy\\sourcefiles\\新生--张晓杰.xls"

    target_file_path = "D:\\workspace_test\\AutoDeploy\\sourcefiles\\四川传媒学院宿舍信息导入样表.xlsx"

    source_excel = xlrd.open_workbook(source_file_path)

    id_name_dics ={}

    source_excel_sheet = source_excel.sheet_by_name("模板表")

    for index in range(1, 147):
        s_id = source_excel_sheet.cell_value(index, 4)
        s_name = source_excel_sheet.cell_value(index, 1)
        id_name_dics[s_name] = s_id
        print(s_id+"-"+s_name)

    target_excel = openpyxl.load_workbook(target_file_path)

    target_excel_sheet = target_excel["学生宿舍信息"]

    for index in range(3, 150):
        temp_name = target_excel_sheet.cell(index, 2).value
        if id_name_dics.get(temp_name):
            target_excel_sheet.cell(index, 1).value = id_name_dics[temp_name]

    target_excel.save("D:\\workspace_test\\AutoDeploy\\sourcefiles\\新生--张晓杰.xlsx")